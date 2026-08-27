#!/usr/bin/env python3
"""Extract the GSE165816 participant/sample map from the source workbook.

This public derivative implements the same row, group, and tissue mapping rules
as the executed JavaScript utility, using the widely available ``openpyxl``
package. It reads Sheet1!A1:R28, refuses to overwrite outputs, and never changes
the source workbook.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


GROUPS = {
    "DFU-NH": "DFU_NONHEALER",
    "DFU-H": "DFU_HEALER",
    "DM": "DIABETES_NO_DFU",
    "H": "HEALTHY_NONDIABETIC",
}

SAMPLE_COLUMNS = (
    (15, "FOOT_SKIN", "P"),
    (16, "FOREARM_SKIN", "Q"),
    (17, "PBMC", "R"),
)

FIELDNAMES = (
    "participant_alias",
    "source_workbook_row",
    "group_source",
    "group_normalized",
    "age_years",
    "sex",
    "sample_id",
    "tissue_context_from_supplement",
    "source_cell",
    "participant_alias_status",
)


def sample_ids(value: object) -> list[str]:
    if value is None or str(value).strip() == "No":
        return []
    match = re.fullmatch(r"Yes \(([^)]+)\)", str(value))
    if match is None:
        raise ValueError(f"Unexpected sample cell: {value}")
    return [item.strip() for item in match.group(1).split("&")]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_workbook", type=Path)
    parser.add_argument("output_csv", type=Path)
    parser.add_argument("output_summary", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_workbook = args.input_workbook.expanduser().resolve()
    output_csv = args.output_csv.expanduser().resolve()
    output_summary = args.output_summary.expanduser().resolve()

    for output in (output_csv, output_summary):
        if output.exists():
            raise FileExistsError(f"Refusing to overwrite existing output: {output}")

    workbook = load_workbook(input_workbook, read_only=True, data_only=True)
    if "Sheet1" not in workbook.sheetnames:
        raise ValueError("Expected worksheet Sheet1")
    worksheet = workbook["Sheet1"]
    values = [list(row) for row in worksheet.iter_rows(min_row=1, max_row=28, min_col=1, max_col=18, values_only=True)]
    workbook.close()

    expected_headers = {
        0: "Group",
        1: "Age (years)",
        2: "Sex",
        15: "Foot Skin (sample ID in GSE165816)",
        16: "Forearm Biopsy (sample ID in GSE165816)",
        17: "PBMCs (sample ID in GSE165816)",
    }
    for index, expected in expected_headers.items():
        if values[0][index] != expected:
            raise ValueError(f"Header mismatch at column {index + 1}: {values[0][index]}")

    rows: list[dict[str, object]] = []
    participant_group_counts: Counter[str] = Counter()
    for row_index, row in enumerate(values[1:], start=1):
        participant_alias = f"GSE165816_P{row_index:03d}"
        group_source = str(row[0])
        if group_source not in GROUPS:
            raise ValueError(f"Unexpected group: {group_source}")
        group_normalized = GROUPS[group_source]
        participant_group_counts[group_normalized] += 1

        for column_index, tissue, excel_column in SAMPLE_COLUMNS:
            for sample_id in sample_ids(row[column_index]):
                rows.append(
                    {
                        "participant_alias": participant_alias,
                        "source_workbook_row": row_index + 1,
                        "group_source": group_source,
                        "group_normalized": group_normalized,
                        "age_years": row[1],
                        "sex": row[2],
                        "sample_id": sample_id,
                        "tissue_context_from_supplement": tissue,
                        "source_cell": f"Sheet1!{excel_column}{row_index + 1}",
                        "participant_alias_status": "DERIVED_ROW_ALIAS_NOT_A_PUBLIC_CLINICAL_IDENTIFIER",
                    }
                )

    unique_sample_ids = {str(row["sample_id"]) for row in rows}
    tissue_counts = Counter(str(row["tissue_context_from_supplement"]) for row in rows)
    summary = {
        "schema_version": "1.0",
        "input_workbook": str(input_workbook),
        "worksheet_range": "Sheet1!A1:R28",
        "participant_count": len(values) - 1,
        "sample_count": len(rows),
        "unique_sample_count": len(unique_sample_ids),
        "participant_group_counts": dict(participant_group_counts),
        "sample_tissue_counts": dict(tissue_counts),
        "validation_status": (
            "PASS"
            if len(values) - 1 == 27 and len(rows) == 54 and len(unique_sample_ids) == 54
            else "FAIL"
        ),
        "interpretation_boundary": (
            "Participant aliases are deterministic row aliases. The workbook establishes which public "
            "sample IDs belong to the same participant but does not publish clinical participant identifiers."
        ),
    }
    if summary["validation_status"] != "PASS":
        raise RuntimeError(f"Mapping validation failed: {summary}")

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    output_summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
